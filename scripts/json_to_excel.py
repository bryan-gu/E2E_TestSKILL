import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json, glob, re, os, argparse

HEADER_FONT = Font(name='微软雅黑', size=12, bold=True)
HEADER_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
DATA_FONT = Font(name='微软雅黑', size=11)
HEADER_ALIGN = Alignment(wrap_text=True, vertical='center')
DATA_ALIGN = Alignment(wrap_text=True, vertical='top')
COL_WIDTHS = [14.875, 16.625, 20.375, 21.125, 17.875, 26.5, 25.875, 8.675, 18.0, 18.0]
HEADERS = ['测试用例ID', '模块', '标题', '前置条件', '测试数据', '测试步骤', '预期结果', '实际结果',
           '覆盖功能点', '关联接口']

def create_sheet(wb, name):
    ws = wb.create_sheet(title=name)
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    return ws

def add_row(ws, row, data):
    keys = ['id', 'module', 'title', 'precondition', 'test_data', 'steps', 'expected']
    for ci, key in enumerate(keys, 1):
        c = ws.cell(row=row, column=ci, value=data.get(key, ''))
        c.font = DATA_FONT
        c.alignment = DATA_ALIGN
    # 第 8 列「实际结果」默认空（保留 V1 特例）
    ws.cell(row=row, column=8, value='')
    # V2 新增第 9 / 10 列：覆盖功能点 / 关联接口
    # 用 '；'.join 把数组拼成单格字符串；老 cases.json 无字段 → .get([]) → 空串（V1 兼容）
    covers_val = '；'.join(data.get('covers', []) or [])
    api_val = '；'.join(data.get('tests_api', []) or [])
    c9 = ws.cell(row=row, column=9, value=covers_val)
    c9.font = DATA_FONT
    c9.alignment = DATA_ALIGN
    c10 = ws.cell(row=row, column=10, value=api_val)
    c10.font = DATA_FONT
    c10.alignment = DATA_ALIGN

def repair_json(filepath):
    """尝试修复常见的 JSON 问题（如字符串内未转义的双引号）"""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    # 先尝试直接解析
    try:
        json.loads(content)
        return content  # 无需修复
    except json.JSONDecodeError:
        pass
    # 逐行修复：将字符串值中未转义的 " 替换为 「」
    lines = content.split('\n')
    fixed_lines = []
    for line in lines:
        # 匹配 "key": "value" 模式，对 value 部分修复
        m = re.match(r'^(\s*"[^"]+"\s*:\s*)"(.*)"(\s*,?\s*)$', line)
        if m:
            prefix, value, suffix = m.group(1), m.group(2), m.group(3)
            # 将 value 中未转义的 " 替换为 「」
            # 已转义的 \" 保留，未转义的 " 替换
            fixed_value = ''
            i = 0
            while i < len(value):
                if value[i] == '\\' and i + 1 < len(value) and value[i+1] == '"':
                    fixed_value += '\\"'
                    i += 2
                elif value[i] == '"':
                    # 未转义的引号，用「」交替替换
                    open_count = fixed_value.count('「') - fixed_value.count('」')
                    fixed_value += '「' if open_count % 2 == 0 else '」'
                    i += 1
                else:
                    fixed_value += value[i]
                    i += 1
            line = f'{prefix}"{fixed_value}"{suffix}'
        fixed_lines.append(line)
    fixed = '\n'.join(fixed_lines)
    # 验证修复结果
    try:
        json.loads(fixed)
        # 修复成功，回写文件
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(fixed)
        return fixed
    except json.JSONDecodeError as e:
        print(f'  警告: {filepath} 修复失败 ({e})，跳过')
        return None

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='将 cases.json 转换为 testCase.xlsx')
    parser.add_argument('--input', '-i', default='需求文档/需求功能点',
                        help='功能点目录路径（默认：需求文档/需求功能点）')
    parser.add_argument('--output', '-o', default='测试用例/testCase.xlsx',
                        help='输出 xlsx 路径（默认：测试用例/testCase.xlsx）')
    parser.add_argument('--sprint', '-s', default=None,
                        help='Sprint 名称（如 sprint0、sprint1），自动设置输入输出路径')
    args = parser.parse_args()

    # 根据 sprint 参数推导路径
    if args.sprint:
        sprint = args.sprint
        input_dir = f'需求文档/{sprint}/需求功能点'
        output_file = f'测试用例/{sprint}/{sprint}_testCase.xlsx'
    else:
        input_dir = args.input
        output_file = args.output

    # 确保输出目录存在
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    # 如果输出文件已存在，加载它（支持按模块覆盖）
    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    json_files = sorted(glob.glob(f'{input_dir}/*/cases.json'))
    total = 0
    for jf in json_files:
        content = repair_json(jf)
        if content is None:
            continue
        cases = json.loads(content)
        if not cases:
            continue
        sheet_name = cases[0]['module'][:31]
        # 如果 Sheet 已存在，删除后重建
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = create_sheet(wb, sheet_name)
        for i, case in enumerate(cases, 2):
            add_row(ws, i, case)
        print(f'{sheet_name}: {len(cases)}条')
        total += len(cases)

    wb.save(output_file)
    print(f'总计: {total}条测试用例 → {output_file}')
