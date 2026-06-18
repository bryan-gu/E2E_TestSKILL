"""
将 Playwright 测试执行结果回写到 testCase.xlsx 的"实际结果"列。

用法:
    python write_results.py <xlsx路径> <Sheet名> <结果JSON>

示例:
    python write_results.py "测试用例/sprint_all/testCase.xlsx" "数据整合" '{"TC_001":"通过","TC_002":"失败-超时"}'

结果JSON格式: {用例ID: 结果文本}
    - 通过: "通过"
    - 失败: "失败-{原因}"
    - 跳过: "跳过-{原因}"
"""
import openpyxl
import json
import sys

def main():
    if len(sys.argv) != 4:
        print('用法: python write_results.py <xlsx路径> <Sheet名> <结果JSON>')
        sys.exit(1)

    filepath = sys.argv[1]
    sheet_name = sys.argv[2]
    results = json.loads(sys.argv[3])

    wb = openpyxl.load_workbook(filepath)

    if sheet_name not in wb.sheetnames:
        print(f'错误: Sheet "{sheet_name}" 不存在。可用: {wb.sheetnames}')
        sys.exit(1)

    ws = wb[sheet_name]

    # 找到"测试用例ID"和"实际结果"列索引
    id_col = None
    result_col = None
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val == '测试用例ID':
            id_col = c
        elif val == '实际结果':
            result_col = c

    if id_col is None or result_col is None:
        print('错误: 未找到"测试用例ID"或"实际结果"列')
        sys.exit(1)

    updated = 0
    for r in range(2, ws.max_row + 1):
        case_id = ws.cell(row=r, column=id_col).value
        if not case_id:
            continue
        case_id = str(case_id).strip()
        if case_id in results:
            ws.cell(row=r, column=result_col).value = results[case_id]
            print(f'  {case_id} -> {results[case_id]}')
            updated += 1

    wb.save(filepath)
    print(f'完成: {updated} 条用例已更新')

if __name__ == '__main__':
    main()
